"""
Provincial Energy Storage Peak-Shaving Analysis (Yearly Cost Version)

Runs storage optimization for each province across years 2025-2035,
using year-specific storage unit costs. Outputs detailed Excel reports
including storage capacity vs. peak-shaving supply curves, target
satisfaction rate parameters, and daily first-satisfy capacity.
"""

import pandas as pd
import numpy as np
import numpy_financial as npf
import json
import os
from pathlib import Path
from shapely.geometry import shape, Point
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import CONFIG, calculate_crf
from utils import get_price

warnings.filterwarnings('ignore')

# Year-specific storage unit costs (Yuan/kWh) extracted from CONFIG
YEAR_COSTS = CONFIG["year_costs"]

# Pre-compute CRF using default cost parameters
CONFIG['CRF'] = calculate_crf(CONFIG['cost']['discount_rate'],
                               CONFIG['cost']['project_lifetime'])


# ============================================================================
# 1. Data Loading Module
# ============================================================================


class DataLoader:
    """Handles all data loading for the yearly analysis."""

    def __init__(self, config):
        self.config = config
        self.province_polygons = {}
        self.station_info = {}
        self.station_province_map = {}
        self.electricity_price = {}
        self.gas_price = {}

    def load_province_boundaries(self):
        """Load province boundary GeoJSON."""
        print("Loading province administrative boundaries...")
        filepath = self.config['paths']['province_geojson']
        with open(filepath, 'r', encoding='utf-8') as f:
            geojson_data = json.load(f)
        for feature in geojson_data['features']:
            props = feature['properties']
            province_name = (props.get('name') or props.get('NAME')
                             or props.get('省份') or props.get('省'))
            if province_name:
                self.province_polygons[province_name] = shape(feature['geometry'])
        print(f"  Loaded {len(self.province_polygons)} province boundaries")
        print(f"  Province list: {list(self.province_polygons.keys())}")

    def load_station_info(self):
        """Load power station GeoJSON (ObjectId, Latitude, Longitude, Capacity__MW_)."""
        print("Loading power station geographic data...")
        filepath = self.config['paths']['station_geojson']
        with open(filepath, 'r', encoding='utf-8') as f:
            geojson_data = json.load(f)
        for feature in geojson_data['features']:
            props = feature['properties']
            object_id = str(props.get('ObjectId', '')).strip()
            if object_id.endswith('.0'):
                object_id = object_id[:-2]
            if not object_id:
                continue
            self.station_info[object_id] = {
                'latitude': props.get('Latitude', 0),
                'longitude': props.get('Longitude', 0),
                'capacity_mw': props.get('Capacity__MW_', 0),
            }
        print(f"  Loaded {len(self.station_info)} station records")

    def match_stations_to_provinces(self):
        """Match stations to provinces using shapely Point-in-polygon."""
        print("Matching stations to provinces...")
        matched_count = 0
        unmatched_stations = []
        for object_id, info in self.station_info.items():
            lat, lon = info['latitude'], info['longitude']
            if lat == 0 or lon == 0:
                unmatched_stations.append(object_id)
                continue
            point = Point(lon, lat)
            matched_province = None
            for province_name, polygon in self.province_polygons.items():
                if polygon.contains(point):
                    matched_province = province_name
                    break
            if matched_province:
                self.station_province_map[object_id] = matched_province
                matched_count += 1
            else:
                unmatched_stations.append(object_id)
        print(f"  Successfully matched {matched_count} stations")
        if unmatched_stations:
            print(f"  Warning: {len(unmatched_stations)} stations unmatched")

    def load_electricity_price(self):
        """Load electricity prices from Excel (fuzzy column matching)."""
        print("Loading electricity price data...")
        filepath = self.config['paths']['electricity_price_xlsx']
        df = pd.read_excel(filepath)
        province_col = price_col = None
        for col in df.columns:
            if '省' in col or '自治区' in col or '直辖市' in col:
                province_col = col
            if '基准价' in col or '电价' in col:
                price_col = col
        if province_col is None or price_col is None:
            print(f"  Error: cannot identify column names, columns: {list(df.columns)}")
            return
        for _, row in df.iterrows():
            province = str(row[province_col]).strip()
            self.electricity_price[province] = (
                float(row[price_col]) if pd.notna(row[price_col]) else 0.0
            )
        print(f"  Loaded electricity prices for {len(self.electricity_price)} provinces")

    def load_gas_price(self):
        """Load gas prices from Excel."""
        print("Loading natural gas price data...")
        file_path = self.config['paths']['gas_price_xlsx']
        if not os.path.exists(file_path):
            print(f"  Error: gas price file not found: {file_path}")
            return
        df = pd.read_excel(file_path)
        for _, row in df.iterrows():
            province = str(row.iloc[0]).strip()
            self.gas_price[province] = (
                float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0.0
            )
        print(f"  Loaded gas prices for {len(self.gas_price)} provinces")

    def load_peak_demand(self):
        """Load peak demand Excel, transpose, add month_day, remove Feb 29."""
        print("Loading peak-shaving demand data...")
        file_path = self.config['paths']['peak_demand_xlsx']
        if not os.path.exists(file_path):
            print(f"  Error: peak demand file not found: {file_path}")
            return None
        print(f"  Reading file: {file_path}")
        df = pd.read_excel(file_path, index_col=0)
        df = df.T
        df.index = pd.to_datetime(df.index)
        df = df.sort_index()
        df['month_day'] = df.index.strftime('%m-%d')
        df = df[df['month_day'] != '02-29']
        print(f"    Province count: {len(df.columns) - 1}")
        print(f"    Date range: {df.index.min()} to {df.index.max()}")
        return df

    def load_curtailment_data(self):
        """Load hourly curtailment CSVs by month, aggregate to daily by province.

        Files: TotalCurt_Hourly_{YEAR}{MM}.csv
        Parse ObjectId, match to province, sum hourly to daily (MM-DD format).
        """
        print("Loading curtailment data...")
        year = self.config.get('YEAR', '2020')
        folder = self.config['paths']['curtailment_folder']
        months = [f'{m:02d}' for m in range(1, 13)]
        province_daily_curtailment = {}

        for month in months:
            file_path = os.path.join(folder, f"TotalCurt_Hourly_{year}{month}.csv")
            if not os.path.exists(file_path):
                print(f"  Warning: {year}-{month} data file not found, skipping")
                continue
            print(f"  Reading {year}-{month} data...")
            df = pd.read_csv(file_path, low_memory=False)
            time_columns = [col for col in df.columns if col != 'ObjectId']

            for _, row in df.iterrows():
                object_id = str(row['ObjectId']).strip()
                if object_id.endswith('.0'):
                    object_id = object_id[:-2]
                province = self.station_province_map.get(object_id)
                if not province:
                    continue
                if province not in province_daily_curtailment:
                    province_daily_curtailment[province] = {}

                for col in time_columns:
                    try:
                        date_str = col[:8]
                        month_day = date_str[4:6] + '-' + date_str[6:8]
                        if month_day == '02-29':
                            continue
                    except Exception:
                        continue
                    curtailment_value = (
                        float(row[col]) if pd.notna(row[col]) else 0.0
                    )
                    if month_day not in province_daily_curtailment[province]:
                        province_daily_curtailment[province][month_day] = 0.0
                    province_daily_curtailment[province][month_day] += curtailment_value

        df_curtailment = pd.DataFrame(province_daily_curtailment).sort_index()
        print(f"  Curtailment aggregation complete: "
              f"{len(df_curtailment.columns)} provinces, "
              f"{len(df_curtailment)} days")
        return df_curtailment

    def get_province_price(self, province_name):
        """Fuzzy match electricity price for province."""
        try:
            return get_price(self.electricity_price, province_name)
        except KeyError:
            print(f"    Warning: electricity price not found for {province_name}, "
                  f"using default 0")
            return 0.0

    def get_province_gas_price(self, province_name):
        """Fuzzy match gas price for province."""
        try:
            return get_price(self.gas_price, province_name)
        except KeyError:
            print(f"    Warning: gas price not found for {province_name}, "
                  f"using default 0")
            return 0.0

    def calculate_gas_unit_cost(self, gas_price_per_m3):
        """Calculate gas generation unit cost (Yuan/kWh)."""
        gas_cfg = self.config['gas']
        gas_power_per_m3 = (gas_cfg['heat_value']
                            * gas_cfg['generation_efficiency'] / 3600)
        return gas_price_per_m3 / gas_power_per_m3 + gas_cfg['generator_cost']


# ============================================================================
# 2. Storage Simulation Module (plain Python version for yearly analysis)
# ============================================================================


class StorageSimulator:
    """Storage system simulator (plain Python, no Numba dependency)."""

    def __init__(self, config):
        self.config = config

    def simulate_storage_single_year(self, daily_demand, daily_curtailment,
                                     storage_capacity_kwh, initial_soc_ratio):
        """Simulate one year of storage operation.

        Handles pandas Series via .iloc and pd.notna checks.

        Returns
        -------
        total_peak_supply : float
        final_soc_ratio : float
        daily_supply_list : list[float]
        daily_spillage_list : list[float]
        """
        stg = self.config['storage']
        charge_eff = stg['charge_eff']
        discharge_eff = stg['discharge_eff']
        soc_min = stg['soc_min']
        soc_max = stg['soc_max']
        num_days = len(daily_demand)

        if storage_capacity_kwh <= 0:
            daily_spillage_list = []
            for i in range(num_days):
                curt = (daily_curtailment.iloc[i]
                        if hasattr(daily_curtailment, 'iloc')
                        else daily_curtailment[i])
                curt = curt if pd.notna(curt) else 0.0
                daily_spillage_list.append(float(curt))
            return 0.0, initial_soc_ratio, [0.0] * num_days, daily_spillage_list

        min_energy = storage_capacity_kwh * soc_min
        max_energy = storage_capacity_kwh * soc_max
        current_energy = max(min_energy,
                             min(max_energy,
                                 storage_capacity_kwh * initial_soc_ratio))

        total_peak_supply = 0.0
        daily_supply_list = []
        daily_spillage_list = []

        for i in range(num_days):
            demand = (daily_demand.iloc[i]
                      if hasattr(daily_demand, 'iloc')
                      else daily_demand[i])
            curtailment = (daily_curtailment.iloc[i]
                           if hasattr(daily_curtailment, 'iloc')
                           else daily_curtailment[i])
            demand = float(demand) if pd.notna(demand) else 0.0
            curtailment = float(curtailment) if pd.notna(curtailment) else 0.0

            # Discharge to meet demand
            available_energy = current_energy - min_energy
            actual_supply = min(available_energy * discharge_eff, demand)
            if actual_supply > 0:
                current_energy -= actual_supply / discharge_eff
            total_peak_supply += actual_supply
            daily_supply_list.append(actual_supply)

            # Charge from curtailment
            available_space = max_energy - current_energy
            actual_charge = min(curtailment * charge_eff, available_space)
            current_energy += actual_charge

            # Spillage: curtailment not absorbed by storage
            if charge_eff > 0:
                charged_curtailment = actual_charge / charge_eff
            else:
                charged_curtailment = 0.0
            spillage = curtailment - charged_curtailment
            daily_spillage_list.append(spillage)

        final_soc_ratio = (current_energy / storage_capacity_kwh
                           if storage_capacity_kwh > 0 else 0.0)
        return (total_peak_supply, final_soc_ratio,
                daily_supply_list, daily_spillage_list)

    def simulate_storage_steady_state(self, daily_demand, daily_curtailment,
                                      storage_capacity_kwh):
        """Iterative steady-state simulation until SOC converges.

        Returns
        -------
        total_peak_supply : float
        converged : bool
        iterations : int
        daily_supply_list : list[float]
        daily_spillage_list : list[float]
        """
        conv = self.config['convergence']
        convergence_threshold = (storage_capacity_kwh
                                 * conv['threshold_ratio'])
        max_iterations = conv['max_iterations']

        if storage_capacity_kwh <= 0:
            num_days = len(daily_demand)
            daily_spillage_list = []
            for i in range(num_days):
                curt = (daily_curtailment.iloc[i]
                        if hasattr(daily_curtailment, 'iloc')
                        else daily_curtailment[i])
                curt = float(curt) if pd.notna(curt) else 0.0
                daily_spillage_list.append(curt)
            return 0.0, True, 0, [0.0] * num_days, daily_spillage_list

        initial_soc_ratio = self.config['storage']['soc_min']
        daily_supply_list = []
        daily_spillage_list = []

        for iteration in range(1, max_iterations + 1):
            (total_peak_supply, final_soc_ratio,
             daily_supply_list, daily_spillage_list) = \
                self.simulate_storage_single_year(
                    daily_demand, daily_curtailment,
                    storage_capacity_kwh, initial_soc_ratio)
            if (abs(final_soc_ratio - initial_soc_ratio)
                    * storage_capacity_kwh < convergence_threshold):
                return (total_peak_supply, True, iteration,
                        daily_supply_list, daily_spillage_list)
            initial_soc_ratio = final_soc_ratio

        return (total_peak_supply, False, max_iterations,
                daily_supply_list, daily_spillage_list)

    def check_should_stop(self, daily_demand, daily_supply_list,
                          daily_spillage_list):
        """Check if further capacity increase would help using prefix sum
        of spillage.

        For each day with unmet demand, check whether cumulative spillage
        from prior days is positive. If yes, more capacity could help.
        If no prior spillage exists for any unsatisfied day, stop.

        Returns
        -------
        bool
            True if the search should stop.
        """
        num_days = len(daily_demand)
        if num_days == 0:
            return True

        # Build prefix sum of daily spillage
        spillage_prefix_sum = [0.0] * (num_days + 1)
        for i in range(num_days):
            spillage_prefix_sum[i + 1] = (spillage_prefix_sum[i]
                                          + daily_spillage_list[i])

        for d in range(num_days):
            demand_val = (daily_demand.iloc[d]
                          if hasattr(daily_demand, 'iloc')
                          else daily_demand[d])
            demand_val = float(demand_val) if pd.notna(demand_val) else 0.0
            if demand_val <= 0:
                continue
            supply_val = daily_supply_list[d]
            if supply_val >= demand_val * 0.9999:
                continue

            # Day d is unsatisfied; check prior cumulative spillage
            if d == 0:
                continue
            history_spillage = spillage_prefix_sum[d]
            if history_spillage > 1e-6:
                return False

        return True

    def calculate_irr(self, initial_cost, annual_revenue):
        """Calculate IRR using numpy_financial.irr."""
        if initial_cost <= 0 or annual_revenue <= 0:
            return None
        years = self.config['cost']['project_lifetime']
        cash_flows = [-initial_cost] + [annual_revenue] * years
        try:
            irr = npf.irr(cash_flows)
            if np.isnan(irr) or np.isinf(irr):
                return None
            return irr * 100
        except Exception:
            return None

    def calculate_net_profit_zero_point(self, results, electricity_price):
        """Find capacity where net profit crosses zero by linear interpolation.

        Returns
        -------
        dict or None
            Dictionary with interpolated values at the zero-profit point.
        """
        if len(results) < 2:
            return None

        cost_cfg = self.config['cost']
        stg = self.config['storage']

        for i in range(1, len(results)):
            p0 = results[i - 1]['annual_net_profit']
            p1 = results[i]['annual_net_profit']
            if p0 is not None and p1 is not None:
                if (p0 > 0 and p1 <= 0) or (p0 >= 0 and p1 < 0):
                    cap1 = results[i - 1]['capacity_mwh']
                    cap2 = results[i]['capacity_mwh']
                    if p0 != p1:
                        zero_cap_mwh = cap1 + (0 - p0) * (cap2 - cap1) / (p1 - p0)
                    else:
                        zero_cap_mwh = cap1

                    zero_cap_kwh = zero_cap_mwh * 1000
                    s0 = results[i - 1]['peak_supply_kwh']
                    s1 = results[i]['peak_supply_kwh']
                    zero_supply = (s0 + (0 - p0) * (s1 - s0) / (p1 - p0)
                                   if p0 != p1 else s0)
                    r0 = results[i - 1]['satisfy_rate_pct']
                    r1 = results[i]['satisfy_rate_pct']
                    zero_rate = (r0 + (0 - p0) * (r1 - r0) / (p1 - p0)
                                 if p0 != p1 else r0)

                    initial_cost = zero_cap_kwh * cost_cfg['storage_cost_per_kwh']
                    annual_cost = initial_cost * self.config['CRF']
                    annual_revenue = zero_supply * electricity_price
                    annual_net_profit = annual_revenue - annual_cost
                    irr = self.calculate_irr(initial_cost, annual_revenue)
                    unit_profit = (annual_net_profit / zero_supply
                                   if zero_supply > 0 else None)
                    discharge_eff = stg['discharge_eff']
                    efc = ((zero_supply / discharge_eff) / zero_cap_kwh
                           if zero_cap_kwh > 0 else None)

                    return {
                        'capacity_mwh': round(zero_cap_mwh, 4),
                        'peak_supply_kwh': round(zero_supply, 2),
                        'satisfy_rate_pct': round(zero_rate, 2),
                        'efc': round(efc, 2) if efc is not None else None,
                        'initial_cost': round(initial_cost, 2),
                        'annual_cost': round(annual_cost, 2),
                        'annual_revenue': round(annual_revenue, 2),
                        'annual_net_profit': round(annual_net_profit, 2),
                        'unit_profit': (round(unit_profit, 4)
                                        if unit_profit is not None else None),
                        'irr_pct': (round(irr, 2)
                                    if irr is not None else None),
                    }
        return None

    def calculate_target_satisfy_rate_point(self, results, target_rate,
                                            electricity_price):
        """Find capacity for a target satisfaction rate by linear interpolation.

        Returns
        -------
        dict or None
            Dictionary with interpolated values at the target rate.
        """
        if len(results) < 2:
            return None

        cost_cfg = self.config['cost']
        stg = self.config['storage']

        for i in range(1, len(results)):
            r0 = results[i - 1]['satisfy_rate_pct']
            r1 = results[i]['satisfy_rate_pct']
            if r0 is not None and r1 is not None:
                if (r0 < target_rate <= r1) or (r0 <= target_rate < r1):
                    cap1 = results[i - 1]['capacity_mwh']
                    cap2 = results[i]['capacity_mwh']
                    target_cap_mwh = (cap1 + (target_rate - r0)
                                      * (cap2 - cap1) / (r1 - r0)
                                      if r0 != r1 else cap1)
                    target_cap_kwh = target_cap_mwh * 1000

                    s0 = results[i - 1]['peak_supply_kwh']
                    s1 = results[i]['peak_supply_kwh']
                    target_supply = (s0 + (target_rate - r0)
                                     * (s1 - s0) / (r1 - r0)
                                     if r0 != r1 else s0)

                    initial_cost = (target_cap_kwh
                                    * cost_cfg['storage_cost_per_kwh'])
                    annual_cost = initial_cost * self.config['CRF']
                    annual_revenue = target_supply * electricity_price
                    annual_net_profit = annual_revenue - annual_cost
                    irr = self.calculate_irr(initial_cost, annual_revenue)
                    unit_profit = (annual_net_profit / target_supply
                                   if target_supply > 0 else None)
                    discharge_eff = stg['discharge_eff']
                    efc = ((target_supply / discharge_eff) / target_cap_kwh
                           if target_cap_kwh > 0 else None)
                    unit_peak_cost = (annual_cost / target_supply
                                      if target_supply > 0 else None)

                    return {
                        'capacity_mwh': round(target_cap_mwh, 4),
                        'peak_supply_kwh': round(target_supply, 2),
                        'satisfy_rate_pct': round(target_rate, 2),
                        'efc': round(efc, 2) if efc is not None else None,
                        'initial_cost': round(initial_cost, 2),
                        'annual_cost': round(annual_cost, 2),
                        'annual_revenue': round(annual_revenue, 2),
                        'annual_net_profit': round(annual_net_profit, 2),
                        'unit_profit': (round(unit_profit, 4)
                                        if unit_profit is not None else None),
                        'irr_pct': (round(irr, 2)
                                    if irr is not None else None),
                        'unit_peak_cost': (round(unit_peak_cost, 4)
                                           if unit_peak_cost is not None
                                           else None),
                    }
        return None

    def search_optimal_storage(self, daily_demand, daily_curtailment,
                               province_name, electricity_price):
        """Search optimal storage by incrementing step_kwh.

        Tracks daily first-satisfy capacity.

        Returns
        -------
        results : list[dict]
        optimal_capacity : float
            Optimal capacity in MWh.
        max_satisfy_rate : float
        is_fully_satisfied : bool
        zero_point_info : dict or None
        daily_first_satisfy_capacity : dict
            Mapping of day_index -> first capacity (MWh) at which
            that day's demand was fully met.
        """
        stg = self.config['storage']
        step_mwh = stg['step']
        step_kwh = step_mwh * 1000
        discharge_eff = stg['discharge_eff']
        total_demand = float(daily_demand.sum())

        if total_demand <= 0:
            print(f"    Warning: {province_name} has zero peak-shaving demand, "
                  f"skipping")
            return [], 0, 0, False, None, {}

        results = []
        optimal_capacity = None
        max_satisfy_rate = 0
        max_peak_supply = 0
        is_fully_satisfied = False
        capacity_kwh = 0

        num_days = len(daily_demand)
        daily_first_satisfy_capacity = {}

        # Days with zero demand are satisfied at capacity 0
        for i in range(num_days):
            demand_val = (daily_demand.iloc[i]
                          if hasattr(daily_demand, 'iloc')
                          else daily_demand[i])
            demand_val = float(demand_val) if pd.notna(demand_val) else 0.0
            if demand_val <= 0:
                daily_first_satisfy_capacity[i] = 0.0

        while True:
            if capacity_kwh == 0:
                peak_supply = 0.0
                daily_supply_list = [0.0] * num_days
                daily_spillage_list = []
                for i in range(num_days):
                    curt = (daily_curtailment.iloc[i]
                            if hasattr(daily_curtailment, 'iloc')
                            else daily_curtailment[i])
                    curt = float(curt) if pd.notna(curt) else 0.0
                    daily_spillage_list.append(curt)
            else:
                (peak_supply, converged, iterations,
                 daily_supply_list, daily_spillage_list) = \
                    self.simulate_storage_steady_state(
                        daily_demand, daily_curtailment, capacity_kwh)

            satisfy_rate = ((peak_supply / total_demand * 100)
                            if total_demand > 0 else 0)
            if satisfy_rate > max_satisfy_rate:
                max_satisfy_rate = satisfy_rate
            if peak_supply > max_peak_supply:
                max_peak_supply = peak_supply

            # Check daily first-satisfy
            if capacity_kwh > 0:
                for i in range(num_days):
                    if i in daily_first_satisfy_capacity:
                        continue
                    demand_val = (daily_demand.iloc[i]
                                  if hasattr(daily_demand, 'iloc')
                                  else daily_demand[i])
                    demand_val = (float(demand_val)
                                  if pd.notna(demand_val) else 0.0)
                    if (demand_val > 0
                            and len(daily_supply_list) > i
                            and daily_supply_list[i] >= demand_val * 0.9999):
                        daily_first_satisfy_capacity[i] = capacity_kwh / 1000

            cost_cfg = self.config['cost']
            initial_cost = capacity_kwh * cost_cfg['storage_cost_per_kwh']
            annual_cost = initial_cost * self.config['CRF']
            annual_revenue = peak_supply * electricity_price
            annual_net_profit = annual_revenue - annual_cost
            irr = self.calculate_irr(initial_cost, annual_revenue)
            unit_profit = (annual_net_profit / peak_supply
                           if peak_supply > 0 else None)
            efc = ((peak_supply / discharge_eff) / capacity_kwh
                   if capacity_kwh > 0 else None)

            results.append({
                'capacity_mwh': capacity_kwh / 1000,
                'total_demand_kwh': round(total_demand, 2),
                'peak_supply_kwh': round(peak_supply, 2),
                'satisfy_rate_pct': round(satisfy_rate, 2),
                'efc': round(efc, 2) if efc is not None else None,
                'initial_cost': round(initial_cost, 2),
                'annual_cost': round(annual_cost, 2),
                'annual_revenue': round(annual_revenue, 2),
                'annual_net_profit': round(annual_net_profit, 2),
                'unit_profit': (round(unit_profit, 4)
                                if unit_profit is not None else None),
                'irr_pct': round(irr, 2) if irr is not None else None,
            })

            # Stop condition 1: demand fully satisfied
            if peak_supply >= total_demand * 0.9999:
                optimal_capacity = capacity_kwh / 1000
                is_fully_satisfied = True
                break

            # Stop condition 2: spillage-based early termination
            if capacity_kwh > 0:
                should_stop = self.check_should_stop(
                    daily_demand, daily_supply_list, daily_spillage_list)
                if should_stop:
                    optimal_capacity = capacity_kwh / 1000
                    is_fully_satisfied = False
                    print(f"    Warning: curtailment exhausted -- all prior "
                          f"spillage absorbed; max satisfy rate "
                          f"{max_satisfy_rate:.2f}%")
                    break

            # Stop condition 3: upper bound safety
            capacity_kwh += step_kwh
            if capacity_kwh > total_demand * 10:
                print(f"    Warning: {province_name} reached search upper "
                      f"bound without full satisfaction")
                optimal_capacity = capacity_kwh / 1000
                is_fully_satisfied = False
                break

        zero_point_info = self.calculate_net_profit_zero_point(
            results, electricity_price)
        return (results, optimal_capacity, max_satisfy_rate,
                is_fully_satisfied, zero_point_info,
                daily_first_satisfy_capacity)


# ============================================================================
# 3. Result Export Module
# ============================================================================


class ResultExporter:
    """Handles exporting analysis results to formatted Excel workbooks."""

    def __init__(self, config):
        self.config = config

    def _style(self):
        """Return standard openpyxl styling objects."""
        header_font = Font(bold=True)
        header_fill = PatternFill('solid', fgColor='CCE5FF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        return header_font, header_fill, header_alignment, thin_border

    def export_detail_excel(self, all_results, output_path):
        """Export per-province storage detail sheets with openpyxl formatting."""
        print("Exporting detailed results to Excel...")
        wb = Workbook()
        wb.remove(wb.active)
        header_font, header_fill, header_alignment, thin_border = self._style()
        column_widths = [18, 18, 18, 16, 14, 18, 16, 18, 16, 20, 12]

        for province, results in all_results.items():
            if not results:
                continue
            ws = wb.create_sheet(title=province[:31])
            df = pd.DataFrame(results)
            # English headers
            header_map = {
                'capacity_mwh': 'Capacity (MWh)',
                'total_demand_kwh': 'Total Demand (kWh)',
                'peak_supply_kwh': 'Peak Supply (kWh)',
                'satisfy_rate_pct': 'Satisfy Rate (%)',
                'efc': 'EFC (cycles/yr)',
                'initial_cost': 'Initial Cost (Yuan)',
                'annual_cost': 'Annual Cost (Yuan)',
                'annual_revenue': 'Annual Revenue (Yuan)',
                'annual_net_profit': 'Annual Net Profit (Yuan)',
                'unit_profit': 'Unit Profit (Yuan/kWh)',
                'irr_pct': 'IRR (%)',
            }
            headers = [header_map.get(c, c) for c in df.columns]

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            for row_idx, row_data in enumerate(df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='right')
            for col_idx, width in enumerate(column_widths, 1):
                if col_idx <= len(headers):
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    ws.column_dimensions[col_letter].width = width

        wb.save(output_path)
        print(f"  Detail results saved: {output_path}")

    def export_summary_excel(self, all_summary_data, all_zero_point_data,
                             output_path):
        """Export summary sheet + zero-point sheet."""
        print("Exporting summary results to Excel...")
        wb = Workbook()
        header_font, header_fill, header_alignment, thin_border = self._style()

        # Sheet 1: Province Summary
        ws1 = wb.active
        ws1.title = "Province Summary"
        headers1 = [
            'Province', 'Total Demand (kWh)', 'Capacity (MWh)',
            'Peak Supply (kWh)', 'Max Satisfy Rate (%)',
            'EFC (cycles/yr)', 'Initial Cost (Yuan)', 'Annual Cost (Yuan)',
            'Annual Revenue (Yuan)', 'Annual Net Profit (Yuan)',
            'Unit Profit (Yuan/kWh)', 'IRR (%)',
        ]
        for col_idx, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        for row_idx, data in enumerate(all_summary_data, 2):
            for col_idx, value in enumerate(data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='left' if col_idx <= 1 else 'right')
        col_widths1 = [15, 20, 16, 18, 20, 14, 18, 16, 16, 16, 20, 12]
        for i, w in enumerate(col_widths1, 1):
            ws1.column_dimensions[
                ws1.cell(row=1, column=i).column_letter].width = w

        # Sheet 2: Net Profit Zero Point
        ws2 = wb.create_sheet(title="Zero Profit Point")
        headers2 = [
            'Province', 'Zero-Point Capacity (MWh)', 'Peak Supply (kWh)',
            'Satisfy Rate (%)', 'EFC (cycles/yr)',
            'Initial Cost (Yuan)', 'Annual Cost (Yuan)',
            'Annual Revenue (Yuan)', 'Annual Net Profit (Yuan)',
            'Unit Profit (Yuan/kWh)', 'IRR (%)',
        ]
        for col_idx, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        for row_idx, data in enumerate(all_zero_point_data, 2):
            for col_idx, value in enumerate(data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='left' if col_idx <= 1 else 'right')
        col_widths2 = [15, 20, 18, 16, 14, 18, 16, 16, 16, 20, 12]
        for i, w in enumerate(col_widths2, 1):
            ws2.column_dimensions[
                ws2.cell(row=1, column=i).column_letter].width = w

        wb.save(output_path)
        print(f"  Summary results saved: {output_path}")

    def export_target_rate_excel(self, all_target_rate_data, output_path):
        """Export target satisfaction rate sheets."""
        print("Exporting target satisfy rate results to Excel...")
        wb = Workbook()
        wb.remove(wb.active)
        header_font, header_fill, header_alignment, thin_border = self._style()
        headers = [
            'Province', 'Capacity (MWh)', 'Peak Supply (kWh)',
            'Satisfy Rate (%)', 'EFC (cycles/yr)',
            'Initial Cost (Yuan)', 'Annual Cost (Yuan)',
            'Annual Revenue (Yuan)', 'Annual Net Profit (Yuan)',
            'Unit Profit (Yuan/kWh)', 'IRR (%)',
            'Unit Peak Cost (Yuan/kWh)', 'Gas Unit Cost (Yuan/kWh)',
            'Storage/Gas Cost Ratio',
        ]
        col_widths = [15, 18, 18, 16, 14, 18, 16, 16, 16, 20, 12, 22, 22, 20]

        for target_rate, province_data_list in all_target_rate_data.items():
            ws = wb.create_sheet(title=f"Rate {target_rate}%")
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            for row_idx, data in enumerate(province_data_list, 2):
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(
                        horizontal='left' if col_idx <= 1 else 'right')
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[
                    ws.cell(row=1, column=i).column_letter].width = w

        wb.save(output_path)
        print(f"  Target rate results saved: {output_path}")

    def export_daily_first_satisfy_excel(self, all_daily_first_satisfy,
                                         month_day_list, daily_demand_dict,
                                         output_path):
        """Export daily first-satisfy capacity sheets."""
        print("Exporting daily first-satisfy capacity to Excel...")
        wb = Workbook()
        wb.remove(wb.active)
        header_font, header_fill, header_alignment, thin_border = self._style()
        headers = [
            'Date (MM-DD)', 'Daily Demand (kWh)',
            'First-Satisfy Capacity (MWh)',
            'Initial Cost (Yuan)', 'Annual Cost (Yuan)',
        ]
        col_widths = [12, 22, 26, 18, 16]

        cost_cfg = self.config['cost']

        for province, daily_first_satisfy_capacity in \
                all_daily_first_satisfy.items():
            ws = wb.create_sheet(title=province[:31])
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            province_demand = daily_demand_dict.get(province, {})

            for row_idx, (day_idx, month_day) in \
                    enumerate(enumerate(month_day_list), 2):
                # Date column
                cell = ws.cell(row=row_idx, column=1, value=month_day)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

                # Daily demand
                demand_val = province_demand.get(day_idx, 0.0)
                cell = ws.cell(row=row_idx, column=2,
                               value=round(demand_val, 2))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')

                # First-satisfy capacity and costs
                if day_idx in daily_first_satisfy_capacity:
                    cap_mwh = daily_first_satisfy_capacity[day_idx]
                    cap_kwh = cap_mwh * 1000
                    initial_cost = (cap_kwh
                                    * cost_cfg['storage_cost_per_kwh'])
                    annual_cost = initial_cost * self.config['CRF']

                    cell = ws.cell(row=row_idx, column=3,
                                   value=round(cap_mwh, 4))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='right')

                    cell = ws.cell(row=row_idx, column=4,
                                   value=round(initial_cost, 2))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='right')

                    cell = ws.cell(row=row_idx, column=5,
                                   value=round(annual_cost, 2))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='right')
                else:
                    for col_idx in range(3, 6):
                        cell = ws.cell(row=row_idx, column=col_idx,
                                       value='Not satisfied')
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')

            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[
                    ws.cell(row=1, column=i).column_letter].width = w

        wb.save(output_path)
        print(f"  Daily first-satisfy data saved: {output_path}")


# ============================================================================
# 4. National Aggregation Functions
# ============================================================================


def calculate_national_summary(all_summary_data, config):
    """Aggregate province summaries to national totals.

    Sum: total_demand, capacity, supply, costs, revenue, profit.
    Recompute: satisfy_rate, EFC, unit_profit, IRR.

    Parameters
    ----------
    all_summary_data : list[list]
        Each row: [province, total_demand, capacity, supply,
                   max_satisfy_rate, efc, initial_cost, annual_cost,
                   annual_revenue, annual_net_profit, unit_profit, irr]
                   Indices:  0        1          2        3
                             4               5    6              7
                             8                9                10   11

    Returns
    -------
    list or None
        National summary row, or None if no data.
    """
    if not all_summary_data:
        return None

    sum_total_demand = 0.0
    sum_capacity = 0.0
    sum_supply = 0.0
    sum_initial_cost = 0.0
    sum_annual_cost = 0.0
    sum_annual_revenue = 0.0
    sum_annual_net_profit = 0.0

    for row in all_summary_data:
        sum_total_demand += row[1] if row[1] is not None else 0.0
        sum_capacity += row[2] if row[2] is not None else 0.0
        sum_supply += row[3] if row[3] is not None else 0.0
        sum_initial_cost += row[6] if row[6] is not None else 0.0
        sum_annual_cost += row[7] if row[7] is not None else 0.0
        sum_annual_revenue += row[8] if row[8] is not None else 0.0
        sum_annual_net_profit += row[9] if row[9] is not None else 0.0

    national_satisfy_rate = (round(sum_supply / sum_total_demand * 100, 2)
                             if sum_total_demand > 0 else None)
    discharge_eff = config['storage']['discharge_eff']
    sum_capacity_kwh = sum_capacity * 1000
    national_efc = (round((sum_supply / discharge_eff) / sum_capacity_kwh, 2)
                    if sum_capacity_kwh > 0 else None)
    national_unit_profit = (round(sum_annual_net_profit / sum_supply, 4)
                            if sum_supply > 0 else None)

    # National IRR from aggregate cash flows
    national_irr = None
    if sum_initial_cost > 0 and sum_annual_revenue > 0:
        years = config['cost']['project_lifetime']
        cash_flows = [-sum_initial_cost] + [sum_annual_revenue] * years
        try:
            irr_val = npf.irr(cash_flows)
            if not (np.isnan(irr_val) or np.isinf(irr_val)):
                national_irr = round(irr_val * 100, 2)
        except Exception:
            pass

    return [
        'National',
        round(sum_total_demand, 2),
        round(sum_capacity, 4),
        round(sum_supply, 2),
        national_satisfy_rate,
        national_efc,
        round(sum_initial_cost, 2),
        round(sum_annual_cost, 2),
        round(sum_annual_revenue, 2),
        round(sum_annual_net_profit, 2),
        national_unit_profit,
        national_irr,
    ]


def calculate_national_zero_point(all_zero_point_data, config):
    """Aggregate province zero-point data to national.

    Parameters
    ----------
    all_zero_point_data : list[list]
        Each row: [province, capacity, supply, satisfy_rate, efc,
                   initial_cost, annual_cost, annual_revenue,
                   annual_net_profit, unit_profit, irr]
                   Indices: 0  1  2  3  4  5  6  7  8  9  10

    Returns
    -------
    list or None
    """
    if not all_zero_point_data:
        return None

    sum_capacity = 0.0
    sum_supply = 0.0
    sum_initial_cost = 0.0
    sum_annual_cost = 0.0
    sum_annual_revenue = 0.0
    sum_annual_net_profit = 0.0
    has_valid = False

    for row in all_zero_point_data:
        if row[1] is None:
            continue
        has_valid = True
        sum_capacity += row[1] if row[1] is not None else 0.0
        sum_supply += row[2] if row[2] is not None else 0.0
        sum_initial_cost += row[5] if row[5] is not None else 0.0
        sum_annual_cost += row[6] if row[6] is not None else 0.0
        sum_annual_revenue += row[7] if row[7] is not None else 0.0
        sum_annual_net_profit += row[8] if row[8] is not None else 0.0

    if not has_valid:
        return ['National'] + [None] * 10

    discharge_eff = config['storage']['discharge_eff']
    sum_capacity_kwh = sum_capacity * 1000
    national_efc = (round((sum_supply / discharge_eff) / sum_capacity_kwh, 2)
                    if sum_capacity_kwh > 0 else None)
    national_unit_profit = (round(sum_annual_net_profit / sum_supply, 4)
                            if sum_supply > 0 else None)

    # Satisfy rate is not meaningful as a sum at the zero point
    national_satisfy_rate = None

    national_irr = None
    if sum_initial_cost > 0 and sum_annual_revenue > 0:
        years = config['cost']['project_lifetime']
        cash_flows = [-sum_initial_cost] + [sum_annual_revenue] * years
        try:
            irr_val = npf.irr(cash_flows)
            if not (np.isnan(irr_val) or np.isinf(irr_val)):
                national_irr = round(irr_val * 100, 2)
        except Exception:
            pass

    return [
        'National',
        round(sum_capacity, 4),
        round(sum_supply, 2),
        national_satisfy_rate,
        national_efc,
        round(sum_initial_cost, 2),
        round(sum_annual_cost, 2),
        round(sum_annual_revenue, 2),
        round(sum_annual_net_profit, 2),
        national_unit_profit,
        national_irr,
    ]


def calculate_national_target_rate(province_data_list, config):
    """Aggregate province target rate data to national.

    Includes weighted gas cost.

    Parameters
    ----------
    province_data_list : list[list]
        Each row: [province, capacity, supply, satisfy_rate, efc,
                   initial_cost, annual_cost, annual_revenue,
                   annual_net_profit, unit_profit, irr,
                   unit_peak_cost, gas_unit_cost, cost_ratio]
                   Indices: 0-13

    Returns
    -------
    list or None
    """
    if not province_data_list:
        return None

    sum_capacity = 0.0
    sum_supply = 0.0
    sum_initial_cost = 0.0
    sum_annual_cost = 0.0
    sum_annual_revenue = 0.0
    sum_annual_net_profit = 0.0
    # Weighted gas cost: weight by peak supply
    weighted_gas_cost_sum = 0.0
    gas_weight_sum = 0.0
    has_valid = False

    for row in province_data_list:
        if row[1] is None:
            continue
        has_valid = True
        supply_val = row[2] if row[2] is not None else 0.0
        sum_capacity += row[1] if row[1] is not None else 0.0
        sum_supply += supply_val
        sum_initial_cost += row[5] if row[5] is not None else 0.0
        sum_annual_cost += row[6] if row[6] is not None else 0.0
        sum_annual_revenue += row[7] if row[7] is not None else 0.0
        sum_annual_net_profit += row[8] if row[8] is not None else 0.0
        # Weighted gas cost
        gas_cost = row[12]
        if gas_cost is not None and supply_val > 0:
            weighted_gas_cost_sum += gas_cost * supply_val
            gas_weight_sum += supply_val

    if not has_valid:
        return ['National'] + [None] * 13

    discharge_eff = config['storage']['discharge_eff']
    sum_capacity_kwh = sum_capacity * 1000
    national_efc = (round((sum_supply / discharge_eff) / sum_capacity_kwh, 2)
                    if sum_capacity_kwh > 0 else None)
    national_unit_profit = (round(sum_annual_net_profit / sum_supply, 4)
                            if sum_supply > 0 else None)
    national_unit_peak_cost = (round(sum_annual_cost / sum_supply, 4)
                               if sum_supply > 0 else None)

    national_satisfy_rate = None

    national_irr = None
    if sum_initial_cost > 0 and sum_annual_revenue > 0:
        years = config['cost']['project_lifetime']
        cash_flows = [-sum_initial_cost] + [sum_annual_revenue] * years
        try:
            irr_val = npf.irr(cash_flows)
            if not (np.isnan(irr_val) or np.isinf(irr_val)):
                national_irr = round(irr_val * 100, 2)
        except Exception:
            pass

    # National weighted gas unit cost
    national_gas_cost = (round(weighted_gas_cost_sum / gas_weight_sum, 4)
                         if gas_weight_sum > 0 else None)
    # Storage-to-gas cost ratio
    national_cost_ratio = (
        round(national_unit_peak_cost / national_gas_cost, 4)
        if (national_unit_peak_cost is not None
            and national_gas_cost is not None
            and national_gas_cost > 0)
        else None
    )

    return [
        'National',
        round(sum_capacity, 4),
        round(sum_supply, 2),
        national_satisfy_rate,
        national_efc,
        round(sum_initial_cost, 2),
        round(sum_annual_cost, 2),
        round(sum_annual_revenue, 2),
        round(sum_annual_net_profit, 2),
        national_unit_profit,
        national_irr,
        national_unit_peak_cost,
        national_gas_cost,
        national_cost_ratio,
    ]


# ============================================================================
# 5. Main Function (single-year analysis using current CONFIG settings)
# ============================================================================


def main():
    """Single-year analysis using current CONFIG settings."""
    cost_cfg = CONFIG['cost']
    stg_cfg = CONFIG['storage']

    print("=" * 70)
    print(f"Storage Peak-Shaving Analysis -- "
          f"Unit Cost: {cost_cfg['storage_cost_per_kwh']:.4f} Yuan/kWh")
    print("=" * 70)
    print(f"\nCurrent configuration:")
    print(f"  Storage unit cost: {cost_cfg['storage_cost_per_kwh']:.4f} Yuan/kWh")
    print(f"  Capacity step: {stg_cfg['step']} MWh")
    print(f"  Charge eff: {stg_cfg['charge_eff'] * 100}%  "
          f"Discharge eff: {stg_cfg['discharge_eff'] * 100}%")
    print(f"  SOC range: {stg_cfg['soc_min'] * 100}% ~ "
          f"{stg_cfg['soc_max'] * 100}%")
    print(f"  Project lifetime: {cost_cfg['project_lifetime']} years  "
          f"Discount rate: {cost_cfg['discount_rate'] * 100}%")
    print(f"  CRF: {CONFIG['CRF']:.4f}")
    print(f"  Target satisfy rates: {CONFIG['target_satisfy_rates']}%")
    print(f"  Output directory: {CONFIG['OUTPUT_DIR']}")
    print(f"  Province count: {len(CONFIG['provinces'])}")

    output_dir = Path(CONFIG['OUTPUT_DIR'])
    output_dir.mkdir(parents=True, exist_ok=True)

    print("\n" + "=" * 70)
    print("Loading data")
    print("=" * 70)

    loader = DataLoader(CONFIG)
    loader.load_province_boundaries()
    loader.load_station_info()
    loader.match_stations_to_provinces()
    loader.load_electricity_price()
    loader.load_gas_price()

    df_peak_demand = loader.load_peak_demand()
    if df_peak_demand is None:
        print("Error: failed to load peak demand data, aborting")
        return

    df_curtailment = loader.load_curtailment_data()

    simulator = StorageSimulator(CONFIG)
    exporter = ResultExporter(CONFIG)

    print("\n" + "=" * 70)
    print("Running storage simulation")
    print("=" * 70)

    province_columns_demand = [col for col in df_peak_demand.columns
                               if col != 'month_day']
    common_provinces = list(
        set(province_columns_demand) & set(df_curtailment.columns))

    # Filter to configured province list
    included_set = set(CONFIG['provinces'])
    common_provinces = [p for p in common_provinces if p in included_set]
    excluded_provinces = included_set - set(common_provinces)
    if excluded_provinces:
        print(f"  Warning: specified provinces not found in data: "
              f"{excluded_provinces}")

    demand_month_days = set(df_peak_demand['month_day'].values)
    curtailment_month_days = set(df_curtailment.index.values)
    common_month_days = sorted(demand_month_days & curtailment_month_days)

    print(f"  Common provinces: {len(common_provinces)}, "
          f"common days: {len(common_month_days)}")
    if len(common_provinces) == 0 or len(common_month_days) == 0:
        print("  Error: no common provinces or dates, aborting")
        return

    df_peak_demand_aligned = df_peak_demand.set_index('month_day').loc[
        common_month_days, common_provinces]
    df_curtailment_aligned = df_curtailment.loc[
        common_month_days, common_provinces]
    print("  Data alignment complete")

    month_day_list = list(common_month_days)

    all_results = {}
    all_summary_data = []
    all_zero_point_data = []
    all_target_rate_data = {rate: []
                            for rate in CONFIG['target_satisfy_rates']}
    all_daily_first_satisfy = {}
    daily_demand_dict = {}

    for idx, province in enumerate(common_provinces, 1):
        print(f"\n  [{idx}/{len(common_provinces)}] Analyzing province: "
              f"{province}")
        daily_demand = df_peak_demand_aligned[province]
        daily_curtailment = df_curtailment_aligned[province]
        total_demand = daily_demand.sum()
        total_curtailment = daily_curtailment.sum()
        electricity_price = loader.get_province_price(province)
        gas_price_per_m3 = loader.get_province_gas_price(province)
        gas_unit_cost = loader.calculate_gas_unit_cost(gas_price_per_m3)

        print(f"    Total demand: {total_demand:,.0f} kWh  "
              f"Curtailment: {total_curtailment:,.0f} kWh")
        print(f"    Elec price: {electricity_price} Yuan/kWh  "
              f"Gas unit cost: {gas_unit_cost:.4f} Yuan/kWh")

        # Record daily demand for export
        daily_demand_dict[province] = {
            i: (float(daily_demand.iloc[i])
                if pd.notna(daily_demand.iloc[i]) else 0.0)
            for i in range(len(daily_demand))
        }

        (results, optimal_capacity, max_satisfy_rate,
         is_fully_satisfied, zero_point_info,
         daily_first_satisfy_capacity) = \
            simulator.search_optimal_storage(
                daily_demand, daily_curtailment,
                province, electricity_price)

        # Record daily first-satisfy data
        all_daily_first_satisfy[province] = daily_first_satisfy_capacity
        satisfied_days = len(daily_first_satisfy_capacity)
        total_days = len(daily_demand)
        print(f"    Daily first-satisfy: {satisfied_days}/{total_days} days")

        if results:
            all_results[province] = results
            optimal_capacity_kwh = (optimal_capacity * 1000
                                    if optimal_capacity else 0)
            initial_cost = (optimal_capacity_kwh
                            * cost_cfg['storage_cost_per_kwh'])
            annual_cost = initial_cost * CONFIG['CRF']
            peak_supply = 0.0
            if optimal_capacity and optimal_capacity > 0:
                peak_supply, _, _, _, _ = \
                    simulator.simulate_storage_steady_state(
                        daily_demand, daily_curtailment,
                        optimal_capacity_kwh)
            annual_revenue = peak_supply * electricity_price
            annual_net_profit = annual_revenue - annual_cost
            irr = simulator.calculate_irr(initial_cost, annual_revenue)
            unit_profit = (annual_net_profit / peak_supply
                           if peak_supply > 0 else None)
            discharge_eff = stg_cfg['discharge_eff']
            efc = ((peak_supply / discharge_eff) / optimal_capacity_kwh
                   if optimal_capacity_kwh > 0 else None)

            all_summary_data.append([
                province, round(total_demand, 2), optimal_capacity,
                round(peak_supply, 2), round(max_satisfy_rate, 2),
                round(efc, 2) if efc is not None else None,
                round(initial_cost, 2), round(annual_cost, 2),
                round(annual_revenue, 2), round(annual_net_profit, 2),
                round(unit_profit, 4) if unit_profit is not None else None,
                round(irr, 2) if irr is not None else None,
            ])

            if zero_point_info:
                all_zero_point_data.append([
                    province,
                    zero_point_info['capacity_mwh'],
                    zero_point_info['peak_supply_kwh'],
                    zero_point_info['satisfy_rate_pct'],
                    zero_point_info['efc'],
                    zero_point_info['initial_cost'],
                    zero_point_info['annual_cost'],
                    zero_point_info['annual_revenue'],
                    zero_point_info['annual_net_profit'],
                    zero_point_info['unit_profit'],
                    zero_point_info['irr_pct'],
                ])
                print(f"    Zero-profit capacity: "
                      f"{zero_point_info['capacity_mwh']} MWh")
            else:
                all_zero_point_data.append([province] + [None] * 10)
                print("    Zero-profit point not found")

            for target_rate in CONFIG['target_satisfy_rates']:
                tp = simulator.calculate_target_satisfy_rate_point(
                    results, target_rate, electricity_price)
                if tp:
                    storage_unit_cost = tp['unit_peak_cost']
                    cost_ratio = (
                        round(storage_unit_cost / gas_unit_cost, 4)
                        if (storage_unit_cost is not None
                            and gas_unit_cost > 0)
                        else None
                    )
                    all_target_rate_data[target_rate].append([
                        province,
                        tp['capacity_mwh'],
                        tp['peak_supply_kwh'],
                        tp['satisfy_rate_pct'],
                        tp['efc'],
                        tp['initial_cost'],
                        tp['annual_cost'],
                        tp['annual_revenue'],
                        tp['annual_net_profit'],
                        tp['unit_profit'],
                        tp['irr_pct'],
                        tp['unit_peak_cost'],
                        round(gas_unit_cost, 4),
                        cost_ratio,
                    ])
                    print(f"    Rate {target_rate}% capacity: "
                          f"{tp['capacity_mwh']} MWh")
                else:
                    all_target_rate_data[target_rate].append(
                        [province] + [None] * 11
                        + [round(gas_unit_cost, 4), None])
                    if max_satisfy_rate < target_rate:
                        print(f"    Rate {target_rate}%: not reached "
                              f"(max {max_satisfy_rate:.2f}%)")

            status = ("Fully satisfied" if is_fully_satisfied
                      else "Curtailment insufficient")
            print(f"    {status}  Optimal capacity: {optimal_capacity} MWh  "
                  f"Max satisfy rate: {max_satisfy_rate:.2f}%")

    # ========================================================================
    # National aggregation
    # ========================================================================
    print("\n" + "=" * 70)
    print("Computing national aggregation")
    print("=" * 70)

    # Province summary -> national summary row
    national_summary = calculate_national_summary(all_summary_data, CONFIG)
    if national_summary:
        all_summary_data.append(national_summary)
        print(f"  Province summary - National row: "
              f"demand={national_summary[1]:,.0f} kWh, "
              f"capacity={national_summary[2]:,.1f} MWh, "
              f"satisfy_rate={national_summary[4]}%, "
              f"IRR={national_summary[11]}%")

    # Zero-point -> national row
    national_zero = calculate_national_zero_point(
        all_zero_point_data, CONFIG)
    if national_zero:
        all_zero_point_data.append(national_zero)
        print(f"  Zero-profit - National row: "
              f"capacity={national_zero[1]} MWh, "
              f"IRR={national_zero[10]}%")

    # Target rate -> national row per sheet
    for target_rate in CONFIG['target_satisfy_rates']:
        national_target = calculate_national_target_rate(
            all_target_rate_data[target_rate], CONFIG)
        if national_target:
            all_target_rate_data[target_rate].append(national_target)
            print(f"  Rate {target_rate}% - National row: "
                  f"capacity={national_target[1]} MWh, "
                  f"unit_peak_cost={national_target[11]}, "
                  f"IRR={national_target[10]}%")

    print("\n" + "=" * 70)
    print("Exporting results")
    print("=" * 70)

    exporter.export_detail_excel(
        all_results,
        str(output_dir / "province_storage_detail.xlsx"))
    exporter.export_summary_excel(
        all_summary_data, all_zero_point_data,
        str(output_dir / CONFIG.get('OUTPUT_SUMMARY_EXCEL',
                                    'province_storage_summary.xlsx')))
    exporter.export_target_rate_excel(
        all_target_rate_data,
        str(output_dir / CONFIG.get('OUTPUT_TARGET_RATE_EXCEL',
                                    'target_satisfy_rate_params.xlsx')))
    exporter.export_daily_first_satisfy_excel(
        all_daily_first_satisfy, month_day_list, daily_demand_dict,
        str(output_dir / CONFIG.get('OUTPUT_DAILY_FIRST_SATISFY_EXCEL',
                                    'daily_first_satisfy_capacity.xlsx')))

    print(f"\nAnalysis complete. Provinces analyzed: {len(common_provinces)}")


# ============================================================================
# 6. Yearly Analysis Loop Entry Point
# ============================================================================


def run_yearly_analysis():
    """Loop over YEAR_COSTS, set storage cost per year, run main() for each.

    Output to BASE_OUTPUT_DIR/{year}/ subdirectories.
    """
    base_dir = CONFIG['paths'].get('BASE_OUTPUT_DIR',
                                   os.path.join('output', 'yearly'))

    print("=" * 70)
    print("Yearly Storage Cost Analysis")
    print(f"   Year range: {min(YEAR_COSTS)} ~ {max(YEAR_COSTS)}")
    print(f"   Base output directory: {base_dir}")
    print(f"   Province count: {len(CONFIG['provinces'])}")
    print("=" * 70)
    print("Year\t\tStorage Unit Cost (Yuan/kWh)")
    for year, cost in YEAR_COSTS.items():
        print(f"  {year}\t\t{cost:.4f}")
    print("=" * 70)

    for year, cost in YEAR_COSTS.items():
        CONFIG['cost']['storage_cost_per_kwh'] = cost
        CONFIG['OUTPUT_DIR'] = os.path.join(base_dir, str(year))
        # Recompute CRF (unchanged unless discount rate / lifetime modified)
        CONFIG['CRF'] = calculate_crf(CONFIG['cost']['discount_rate'],
                                       CONFIG['cost']['project_lifetime'])

        print(f"\n{'=' * 70}")
        print(f"Starting year {year}  "
              f"(storage unit cost {cost:.4f} Yuan/kWh)")
        print(f"{'=' * 70}")

        main()

    print(f"\n{'=' * 70}")
    print(f"All years complete. Results saved to: {base_dir}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    run_yearly_analysis()
